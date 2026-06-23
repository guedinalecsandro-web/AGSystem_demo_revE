import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import shutil
import tempfile
from io import BytesIO
from PIL import Image
import base64
import html
import json
import hashlib
import secrets
import hmac
import re
import urllib.parse

# =========================================================
# CONFIGURAÇÃO GERAL
# =========================================================
st.set_page_config(
    page_title="Dashboard PMO",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

CORES_STATUS = {
    "Concluida": "#16A34A",
    "Em Andamento": "#FACC15",
    "Em Atraso": "#DC2626",
    "Concluida atrasada": "#F97316",
    "Sem data alvo": "#6B7280",
}

STATUS_CURTO = {
    "Concluida": "No Prazo",
    "Em Andamento": "Em Atenção",
    "Em Atraso": "Atrasadas",
    "Concluida atrasada": "Concl. Atrasada",
    "Sem data alvo": "Sem Data",
}

# Ajuste aqui caso sua planilha mude de posição/colunas
ABA_PADRAO = "Cronogrma Macro"
LINHA_INICIAL = 15
LINHA_FINAL = 239
COL_ITEM = "C"
COL_ATIVIDADE = "D"
COL_RESPONSAVEL = "P"
COL_DATA_ALVO = "R"
COL_DATA_CONCLUSAO = "S"
COL_DEPENDENCIA = "Depende da Entrega da Atividade (Item)"

# Logo fixo na Sidebar: mantenha o arquivo de imagem na mesma pasta deste programa.
LOGO_SIDEBAR = Path(__file__).with_name("AGSolution_logo.png")
VERSAO_DASHBOARD = "Rev0x_Demo"

# Base local para cronogramas criados diretamente na interface.
CRONOGRAMAS_DIR = Path(__file__).with_name("cronogramas")
DOCUMENTOS_ENTREGA_DIR = Path(__file__).with_name("documentos_entregas")
CRONOGRAMA_ATIVO = CRONOGRAMAS_DIR / "cronograma_ativo.xlsx"
EXTENSAO_CRONOGRAMA_SISTEMA = ".xlsx"

# =========================================================
# CONTROLE DE ACESSO - LOGIN E USUÁRIOS
# =========================================================
USUARIOS_JSON = Path(__file__).with_name("usuarios.json")
USUARIO_ADMIN_PADRAO = "admin"
SENHA_ADMIN_PADRAO = "admin123"


def gerar_hash_senha(senha: str, salt=None) -> str:
    """Gera hash seguro da senha usando PBKDF2."""
    if salt is None:
        salt = secrets.token_hex(16)
    hash_senha = hashlib.pbkdf2_hmac(
        "sha256",
        senha.encode("utf-8"),
        salt.encode("utf-8"),
        120000,
    ).hex()
    return f"{salt}${hash_senha}"


def validar_senha(senha: str, senha_hash: str) -> bool:
    try:
        salt, hash_salvo = senha_hash.split("$", 1)
        hash_digitado = hashlib.pbkdf2_hmac(
            "sha256",
            senha.encode("utf-8"),
            salt.encode("utf-8"),
            120000,
        ).hex()
        return hmac.compare_digest(hash_digitado, hash_salvo)
    except Exception:
        return False


def carregar_usuarios() -> dict:
    """Carrega usuários do arquivo local. Cria admin padrão no primeiro uso."""
    if not USUARIOS_JSON.exists():
        usuarios_iniciais = {
            USUARIO_ADMIN_PADRAO: {
                "nome": "Administrador",
                "usuario": USUARIO_ADMIN_PADRAO,
                "senha_hash": gerar_hash_senha(SENHA_ADMIN_PADRAO),
                "perfil": "Administrador",
                "ativo": True,
                "criado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }
        }
        salvar_usuarios(usuarios_iniciais)
        return usuarios_iniciais

    try:
        return json.loads(USUARIOS_JSON.read_text(encoding="utf-8"))
    except Exception:
        st.error("Erro ao ler usuarios.json. Verifique se o arquivo não está corrompido.")
        st.stop()


def salvar_usuarios(usuarios: dict) -> None:
    USUARIOS_JSON.write_text(
        json.dumps(usuarios, ensure_ascii=False, indent=4),
        encoding="utf-8",
    )


def fazer_logout() -> None:
    for chave in ["autenticado", "usuario_logado", "perfil_logado", "nome_logado"]:
        st.session_state.pop(chave, None)
    st.rerun()


def tela_login() -> None:
    """Tela de login corporativa AGSolution, sem deslocamento vertical e sem campos vazios."""
    logo_b64 = ""
    if LOGO_SIDEBAR.exists():
        try:
            logo_b64 = base64.b64encode(LOGO_SIDEBAR.read_bytes()).decode("utf-8")
        except Exception:
            logo_b64 = ""

    logo_html = (
        f"<img class='ag-login-logo' src='data:image/png;base64,{logo_b64}' alt='AGSolution'>"
        if logo_b64
        else "<div class='ag-login-logo-text'>AGSolution</div>"
    )

    st.markdown(
        """
        <style>
        .stApp {
            background:
                radial-gradient(circle at 18% 18%, rgba(0, 212, 255, 0.20), transparent 28%),
                radial-gradient(circle at 82% 12%, rgba(37, 99, 235, 0.16), transparent 30%),
                linear-gradient(135deg, #020617 0%, #07152E 45%, #020617 100%) !important;
        }

        .block-container {
            padding-top: 1.0rem !important;
            padding-bottom: 1.0rem !important;
            max-width: 1320px !important;
        }

        header[data-testid="stHeader"] {
            background: transparent !important;
            height: 0 !important;
        }

        div[data-testid="stToolbar"],
        div[data-testid="stDecoration"],
        div[data-testid="stStatusWidget"] {
            display: none !important;
        }

        .ag-login-brand {
            background: rgba(2, 6, 23, 0.62);
            border: 1px solid rgba(56, 189, 248, 0.24);
            border-radius: 28px;
            padding: 28px 34px;
            box-shadow: 0 28px 80px rgba(0, 0, 0, 0.35);
            min-height: 500px;
            overflow: hidden;
            position: relative;
        }

        .ag-login-brand:before {
            content: "";
            position: absolute;
            inset: -120px -160px auto auto;
            width: 420px;
            height: 420px;
            border: 1px solid rgba(0, 212, 255, 0.20);
            border-radius: 999px;
            box-shadow: 0 0 70px rgba(0, 212, 255, 0.12);
        }

        .ag-login-logo {
            width: 100%;
            max-width: 500px;
            border-radius: 18px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.32);
            margin-bottom: 20px;
            position: relative;
            z-index: 1;
        }

        .ag-login-logo-text {
            font-size: 58px;
            font-weight: 900;
            color: #E0F2FE;
            letter-spacing: -1px;
            margin-bottom: 18px;
            position: relative;
            z-index: 1;
        }

        .ag-login-kicker {
            color: #38BDF8;
            font-size: 13px;
            font-weight: 800;
            letter-spacing: 0.18em;
            text-transform: uppercase;
            margin-bottom: 12px;
            position: relative;
            z-index: 1;
        }

        .ag-login-title {
            color: #FFFFFF;
            font-size: 36px;
            font-weight: 900;
            line-height: 1.08;
            margin-bottom: 14px;
            position: relative;
            z-index: 1;
        }

        .ag-login-subtitle {
            color: #CBD5E1;
            font-size: 17px;
            line-height: 1.55;
            max-width: 500px;
            margin-bottom: 24px;
            position: relative;
            z-index: 1;
        }

        .ag-feature {
            display: flex;
            gap: 12px;
            align-items: flex-start;
            color: #E2E8F0;
            margin: 12px 0;
            font-size: 15px;
            position: relative;
            z-index: 1;
        }

        .ag-feature-icon {
            width: 34px;
            height: 34px;
            min-width: 34px;
            border-radius: 12px;
            background: rgba(14, 165, 233, 0.16);
            border: 1px solid rgba(56, 189, 248, 0.30);
            display: flex;
            align-items: center;
            justify-content: center;
        }

        .ag-login-panel-header {
            background: rgba(2, 6, 23, 0.62);
            border: 1px solid rgba(56, 189, 248, 0.18);
            border-radius: 28px 28px 0 0;
            padding: 28px 34px 12px 34px;
            margin-top: 0 !important;
        }

        .ag-lock {
            width: 56px;
            height: 56px;
            border-radius: 18px;
            background: linear-gradient(135deg, #0F172A, #1D4ED8);
            color: white;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 26px;
            margin-bottom: 16px;
        }

        .ag-card-title {
            color: #FFFFFF;
            font-size: 30px;
            font-weight: 900;
            line-height: 1.15;
            margin-bottom: 8px;
        }

        .ag-card-subtitle {
            color: #CBD5E1;
            font-size: 15px;
            margin-bottom: 0;
        }

        div[data-testid="stForm"] {
            background: rgba(2, 6, 23, 0.62) !important;
            border: 1px solid rgba(56, 189, 248, 0.18) !important;
            border-top: none !important;
            border-radius: 0 0 28px 28px !important;
            padding: 12px 34px 30px 34px !important;
            margin-top: -1px !important;
            box-shadow: 0 28px 80px rgba(0, 0, 0, 0.35) !important;
        }

        div[data-testid="stForm"] > div {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
        }

        div[data-testid="stForm"] label,
        div[data-testid="stForm"] label p {
            color: #E2E8F0 !important;
            font-weight: 800 !important;
            font-size: 14px !important;
        }

        div[data-testid="stForm"] div[data-testid="stTextInput"] input {
            background: #F8FAFC !important;
            border: 1px solid #CBD5E1 !important;
            border-radius: 12px !important;
            color: #0F172A !important;
            height: 48px !important;
            box-shadow: none !important;
        }

        div[data-testid="stForm"] div[data-testid="stTextInput"] input:focus {
            border: 1px solid #38BDF8 !important;
            box-shadow: 0 0 0 3px rgba(56, 189, 248, 0.20) !important;
        }

        div[data-testid="stForm"] .stCheckbox label,
        div[data-testid="stForm"] .stCheckbox span,
        div[data-testid="stForm"] .stCheckbox p {
            color: #CBD5E1 !important;
        }

        div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] button {
            background: linear-gradient(135deg, #0F172A, #1D4ED8) !important;
            color: #FFFFFF !important;
            border: none !important;
            border-radius: 12px !important;
            height: 48px !important;
            font-weight: 900 !important;
            box-shadow: 0 12px 30px rgba(29, 78, 216, 0.30) !important;
        }

        div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] button:hover {
            filter: brightness(1.10);
        }

        @media (max-width: 960px) {
            .block-container {padding-top: 0.5rem !important;}
            .ag-login-brand {min-height: auto; padding: 24px;}
            .ag-login-title {font-size: 30px;}
            .ag-login-panel-header {padding: 24px 26px 10px 26px;}
            div[data-testid="stForm"] {padding: 10px 26px 26px 26px !important;}
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Layout real do Streamlit, sem abrir <div> antes dos widgets.
    # A versão anterior criava um <div class='ag-login-wrapper'> vazio com altura de tela,
    # causando o deslocamento para baixo e o retângulo branco sem conteúdo.
    col_brand, col_form = st.columns([1.05, 0.95], gap="large")

    with col_brand:
        st.markdown(
            f"""
            <div class='ag-login-brand'>
                {logo_html}
                <div class='ag-login-kicker'>Sistema Corporativo PMO</div>
                <div class='ag-login-title'>Bem-vindo ao<br>Dashboard PMO</div>
                <div class='ag-login-subtitle'>
                    Plataforma para gestão de projetos, cronogramas, indicadores, riscos e relatórios executivos.
                </div>
                <div class='ag-feature'><div class='ag-feature-icon'>📊</div><div><b>Indicadores em tempo real</b><br>Acompanhe desempenho, atrasos e conclusão dos projetos.</div></div>
                <div class='ag-feature'><div class='ag-feature-icon'>📅</div><div><b>Gestão de cronogramas</b><br>Planeje, edite e monitore atividades do projeto.</div></div>
                <div class='ag-feature'><div class='ag-feature-icon'>⚠️</div><div><b>Análise de riscos</b><br>Identifique pendências críticas e pontos de atenção.</div></div>
                <div class='ag-feature'><div class='ag-feature-icon'>⬇️</div><div><b>Relatórios e exportações</b><br>Exporte dados em Excel e CSV para análise gerencial.</div></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_form:
        st.markdown(
            """
            <div class='ag-login-panel-header'>
                <div class='ag-lock'>🔐</div>
                <div class='ag-card-title'>Acesso ao Dashboard PMO</div>
                <div class='ag-card-subtitle'>Informe usuário e senha para acessar o sistema.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.form("form_login"):
            usuario = st.text_input("Usuário", placeholder="Digite seu usuário")
            senha = st.text_input("Senha", type="password", placeholder="Digite sua senha")
            lembrar = st.checkbox("Lembrar-me", value=False)
            entrar = st.form_submit_button("↪ Entrar", width="stretch")

        if entrar:
            usuarios = carregar_usuarios()
            dados_usuario = usuarios.get(usuario.strip())
            if (
                dados_usuario
                and dados_usuario.get("ativo", True)
                and validar_senha(senha, dados_usuario.get("senha_hash", ""))
            ):
                st.session_state["autenticado"] = True
                st.session_state["usuario_logado"] = usuario.strip()
                st.session_state["perfil_logado"] = dados_usuario.get("perfil", "Usuário")
                st.session_state["nome_logado"] = dados_usuario.get("nome", usuario.strip())
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos, ou usuário inativo.")

def exigir_login() -> None:
    if not st.session_state.get("autenticado"):
        tela_login()
        st.stop()


def usuario_eh_admin() -> bool:
    return st.session_state.get("perfil_logado") == "Administrador"


def pagina_administrar_usuarios() -> None:
    st.markdown("<div class='section-title'>Administração de Usuários</div>", unsafe_allow_html=True)
    if not usuario_eh_admin():
        st.warning("Apenas usuário Administrador pode acessar esta área.")
        return

    usuarios = carregar_usuarios()

    st.subheader("Criar novo usuário")
    with st.form("form_novo_usuario"):
        c1, c2 = st.columns(2)
        with c1:
            novo_nome = st.text_input("Nome")
            novo_usuario = st.text_input("Usuário de login")
        with c2:
            novo_perfil = st.selectbox("Perfil", ["Usuário", "Administrador"])
            novo_ativo = st.checkbox("Usuário ativo", value=True)
        nova_senha = st.text_input("Senha inicial", type="password")
        criar = st.form_submit_button("Criar usuário", width="stretch")

    if criar:
        novo_usuario = novo_usuario.strip()
        if not novo_nome.strip() or not novo_usuario or not nova_senha:
            st.error("Preencha nome, usuário e senha.")
        elif novo_usuario in usuarios:
            st.error("Este usuário já existe.")
        else:
            usuarios[novo_usuario] = {
                "nome": novo_nome.strip(),
                "usuario": novo_usuario,
                "senha_hash": gerar_hash_senha(nova_senha),
                "perfil": novo_perfil,
                "ativo": novo_ativo,
                "criado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }
            salvar_usuarios(usuarios)
            st.success("Usuário criado com sucesso.")
            st.rerun()

    st.divider()
    st.subheader("Usuários cadastrados")
    tabela = []
    for u, d in usuarios.items():
        tabela.append({
            "Usuário": u,
            "Nome": d.get("nome", ""),
            "Perfil": d.get("perfil", "Usuário"),
            "Ativo": "Sim" if d.get("ativo", True) else "Não",
            "Criado em": d.get("criado_em", ""),
        })
    st.dataframe(pd.DataFrame(tabela), width="stretch", hide_index=True)

    st.divider()
    st.subheader("Alterar usuário existente")
    usuario_sel = st.selectbox("Selecione o usuário", sorted(usuarios.keys()))
    dados = usuarios[usuario_sel]

    with st.form("form_editar_usuario"):
        edit_nome = st.text_input("Nome", value=dados.get("nome", ""))
        edit_perfil = st.selectbox(
            "Perfil",
            ["Usuário", "Administrador"],
            index=0 if dados.get("perfil", "Usuário") == "Usuário" else 1,
        )
        edit_ativo = st.checkbox("Ativo", value=dados.get("ativo", True))
        edit_senha = st.text_input("Nova senha (deixe vazio para manter)", type="password")
        col_a, col_b = st.columns(2)
        with col_a:
            salvar = st.form_submit_button("Salvar alterações", width="stretch")
        with col_b:
            excluir = st.form_submit_button("Excluir usuário", width="stretch")

    if salvar:
        usuarios[usuario_sel]["nome"] = edit_nome.strip()
        usuarios[usuario_sel]["perfil"] = edit_perfil
        usuarios[usuario_sel]["ativo"] = edit_ativo
        if edit_senha:
            usuarios[usuario_sel]["senha_hash"] = gerar_hash_senha(edit_senha)
        salvar_usuarios(usuarios)
        st.success("Usuário atualizado com sucesso.")
        st.rerun()

    if excluir:
        if usuario_sel == st.session_state.get("usuario_logado"):
            st.error("Você não pode excluir o próprio usuário logado.")
        elif usuario_sel == USUARIO_ADMIN_PADRAO:
            st.error("O usuário admin padrão não pode ser excluído. Você pode alterar a senha ou desativar depois de criar outro administrador.")
        else:
            usuarios.pop(usuario_sel, None)
            salvar_usuarios(usuarios)
            st.success("Usuário excluído com sucesso.")
            st.rerun()


# =========================================================
# CSS - VISUAL ESTILO DASHBOARD WEB
# =========================================================
st.markdown(
    """
    <style>
    .main {background-color: #F5F7FB;}
    section[data-testid="stSidebar"] {background: linear-gradient(180deg, #061B3A 0%, #082B57 100%);}
    section[data-testid="stSidebar"] * {color: white !important;}
    section[data-testid="stSidebar"] .block-container {padding-top: 0.45rem; padding-left: 1rem; padding-right: 1rem;}
    .block-container {padding-top: 2.6rem; padding-bottom: 1rem;}
    .card {
        background: white;
        border: 1px solid #E5E7EB;
        border-radius: 16px;
        padding: 18px 20px;
        box-shadow: 0 6px 20px rgba(15, 23, 42, 0.05);
        min-height: 112px;
    }
    .metric-label {font-size: 13px; color: #475569; font-weight: 600;}
    .metric-value {font-size: 31px; color: #0F172A; font-weight: 800; line-height: 1.1;}
    .metric-delta {font-size: 13px; font-weight: 700; margin-top: 4px;}
    .dashboard-header {
        background: transparent;
        padding: 8px 0 12px 0;
        margin-top: 6px;
        margin-bottom: 8px;
        overflow: visible;
    }
    .header-title {
        font-size: 34px;
        font-weight: 850;
        color: #0F172A;
        margin: 0;
        padding: 4px 0 2px 0;
        line-height: 1.25;
        overflow: visible;
    }
    .header-subtitle {
        font-size: 16px;
        color: #475569;
        margin: 0;
        padding-top: 2px;
        line-height: 1.35;
    }
    .logo-box {
        display: flex;
        align-items: center;
        justify-content: flex-start;
        min-height: 74px;
        padding-top: 2px;
    }
    .client-logo-space {
        min-height: 96px;
        display: flex;
        align-items: center;
        justify-content: flex-start;
        padding-top: 4px;
    }

    .sidebar-logo-box {
        width: 100%;
        display: flex;
        justify-content: flex-start;
        align-items: center;
        padding: 0 0 8px 0;
        margin: 0 0 8px 0;
    }
    .sidebar-caption {
        font-size: 12px;
        color: #D8E6FF;
        line-height: 1.25;
        margin-top: -4px;
        margin-bottom: 14px;
    }
    .executive-header {
        background: #FFFFFF;
        border: 1px solid #E5E7EB;
        border-radius: 18px;
        box-shadow: 0 6px 20px rgba(15, 23, 42, 0.06);
        padding: 18px 22px;
        margin: 2px 0 18px 0;
        display: flex;
        align-items: center;
        gap: 20px;
        min-height: 118px;
        overflow: visible;
    }
    .executive-logo-area {
        width: 145px;
        min-width: 145px;
        display: flex;
        align-items: center;
        justify-content: center;
        border-right: 1px solid #E5E7EB;
        padding-right: 16px;
        min-height: 86px;
    }
    .executive-logo-area img {
        max-width: 125px;
        max-height: 78px;
        object-fit: contain;
    }
    .executive-logo-placeholder {
        width: 118px;
        height: 68px;
        border: 1px dashed #CBD5E1;
        border-radius: 12px;
        color: #94A3B8;
        font-size: 12px;
        display: flex;
        align-items: center;
        justify-content: center;
        text-align: center;
        padding: 6px;
    }
    .executive-main-area {
        flex: 1;
        min-width: 0;
    }
    .executive-title-row {
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        gap: 16px;
    }
    .executive-title {
        font-size: 36px;
        font-weight: 850;
        color: #0F172A;
        line-height: 1.18;
        margin: 0;
        padding: 0;
    }
    .executive-subtitle {
        font-size: 15px;
        color: #475569;
        margin-top: 2px;
        line-height: 1.35;
    }
    .executive-version {
        background: #EFF6FF;
        color: #1D4ED8;
        border: 1px solid #BFDBFE;
        border-radius: 999px;
        padding: 6px 14px;
        font-size: 14px;
        font-weight: 800;
        white-space: nowrap;
    }
    .executive-meta {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 6px 22px;
        margin-top: 10px;
        color: #475569;
        font-size: 13.5px;
        line-height: 1.35;
    }
    .executive-meta b {color: #0F172A;}
    @media (max-width: 900px) {
        .executive-header {align-items: flex-start; gap: 12px; padding: 14px;}
        .executive-logo-area {width: 105px; min-width: 105px; padding-right: 10px;}
        .executive-title {font-size: 28px;}
        .executive-meta {grid-template-columns: 1fr;}
    }
    .section-title {font-size: 19px; font-weight: 800; color:#0F172A; margin: 0 0 10px 0;}
    div[data-testid="stMetric"] {background: white; border-radius: 14px; padding: 14px; border: 1px solid #E5E7EB;}


    /* =======================================================
       SIDEBAR CORPORATIVA PMO - CAMPOS, UPLOADS E BOTÕES
    ======================================================= */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg,#061B3A 0%,#082B57 100%) !important;
    }

    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] div {
        color: #FFFFFF !important;
    }

    section[data-testid="stSidebar"] div[data-testid="stTextInput"] input {
        background-color: #0F3D73 !important;
        color: #FFFFFF !important;
        border: 1px solid #4A90E2 !important;
        border-radius: 10px !important;
        font-weight: 500 !important;
    }

    section[data-testid="stSidebar"] div[data-testid="stTextInput"] input::placeholder {
        color: #D0E3FF !important;
    }

    section[data-testid="stSidebar"] [data-testid="stFileUploader"] {
        background-color: #0F3D73 !important;
        border: 1px solid #4A90E2 !important;
        border-radius: 12px !important;
        padding: 12px !important;
    }

    section[data-testid="stSidebar"] [data-testid="stFileUploader"] section {
        background-color: #0F3D73 !important;
        border: 1px dashed #4A90E2 !important;
        border-radius: 10px !important;
    }

    section[data-testid="stSidebar"] [data-testid="stFileUploader"] * {
        color: #FFFFFF !important;
    }

    section[data-testid="stSidebar"] button {
        background-color: #2563EB !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        transition: 0.3s ease-in-out !important;
    }

    section[data-testid="stSidebar"] button:hover {
        background-color: #1D4ED8 !important;
        color: #FFFFFF !important;
    }

    section[data-testid="stSidebar"] button[kind="secondary"] {
        background-color: #2563EB !important;
        color: #FFFFFF !important;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] {
        background-color: #0F3D73 !important;
        border: 1px solid #4A90E2 !important;
        border-radius: 12px !important;
        padding: 10px !important;
    }

    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.15) !important;
    }

    .sidebar-caption {
        color: #D0E3FF !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Exige autenticação antes de carregar o dashboard
exigir_login()

# =========================================================
# FUNÇÕES
# =========================================================
def carregar_workbook(arquivo_origem):
    try:
        temp_dir = Path(tempfile.gettempdir())
        arquivo_temp = temp_dir / f"pmo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        if hasattr(arquivo_origem, "read"):
            arquivo_temp.write_bytes(arquivo_origem.getvalue())
        else:
            shutil.copy2(Path(arquivo_origem), arquivo_temp)

        return load_workbook(arquivo_temp, data_only=True)

    except PermissionError:
        st.error("Feche o arquivo Excel antes de executar. Ele está bloqueado pelo Excel ou OneDrive.")
        st.stop()
    except FileNotFoundError:
        st.error(f"Arquivo não encontrado: {arquivo_origem}")
        st.stop()
    except Exception as erro:
        st.error(f"Erro ao carregar a planilha: {erro}")
        st.stop()


def normalizar_data(valor):
    return pd.to_datetime(valor, dayfirst=True, errors="coerce")


def calcular_status(row):
    data_alvo = row["Data Alvo"]
    data_atividade = row["Data Entrega"]
    hoje = pd.Timestamp(datetime.now().date())

    if pd.notna(data_atividade) and pd.notna(data_alvo) and data_atividade > data_alvo:
        return "Concluida atrasada"
    if pd.notna(data_atividade) and pd.notna(data_alvo) and data_atividade <= data_alvo:
        return "Concluida"
    if pd.isna(data_atividade) and pd.notna(data_alvo) and data_alvo < hoje:
        return "Em Atraso"
    if pd.isna(data_atividade) and pd.notna(data_alvo) and data_alvo >= hoje:
        return "Em Andamento"
    return "Sem data alvo"


def extrair_dados(wb, aba):
    if aba not in wb.sheetnames:
        st.error(f"Aba '{aba}' não encontrada. Abas disponíveis: {', '.join(wb.sheetnames)}")
        st.stop()

    ws = wb[aba]
    dados = []

    for linha in range(LINHA_INICIAL, LINHA_FINAL):
        item = ws[f"{COL_ITEM}{linha}"].value
        atividade = ws[f"{COL_ATIVIDADE}{linha}"].value

        if atividade is None:
            continue

        alvo = ws[f"{COL_DATA_ALVO}{linha}"].value
        data_atividade = ws[f"{COL_DATA_CONCLUSAO}{linha}"].value

        responsavel = ws[f"{COL_RESPONSAVEL}{linha}"].value
        fase = ws[f"B{linha}"].value if ws.max_column >= 2 else None

        dados.append({
            "Linha": linha,
            "Item": item,
            "Fase": fase if fase else "Não informado",
            "Processo": "",
            "Atividade": atividade,
            "Responsável": responsavel if responsavel else "Não informado",
            "E-mail Responsável": "",
            "Data Início": pd.NaT,
            "Data Alvo": alvo,
            "Data Entrega": data_atividade,
            COL_DEPENDENCIA: "",
            "Observações": "",
        })

    df = pd.DataFrame(dados)
    if df.empty:
        st.warning("Nenhuma atividade encontrada na faixa configurada.")
        st.stop()

    # Normaliza colunas de texto para evitar erro: TypeError '<' not supported between float and str
    for col in ["Fase", "Processo", "Responsável", "E-mail Responsável", "Atividade", "Item", COL_DEPENDENCIA, "Observações"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: limpar_texto(x, padrao="Não informado"))

    # Conversão robusta das datas: evita erro ao subtrair numpy.ndarray/object de Timestamp.
    # Trata datas do Excel, textos, células vazias e valores inválidos como NaT.
    df["Data Início"] = pd.to_datetime(df["Data Início"], errors="coerce", dayfirst=True)
    df["Data Alvo"] = pd.to_datetime(df["Data Alvo"], errors="coerce", dayfirst=True)
    df["Data Entrega"] = pd.to_datetime(df["Data Entrega"], errors="coerce", dayfirst=True)

    df["Status"] = df.apply(calcular_status, axis=1)

    hoje = pd.Timestamp.now().normalize()
    df["Dias para Vencer"] = pd.NA
    mask_data_alvo = df["Data Alvo"].notna()
    df.loc[mask_data_alvo, "Dias para Vencer"] = (
        df.loc[mask_data_alvo, "Data Alvo"] - hoje
    ).dt.days.astype("Int64")

    df["Mês Alvo"] = df["Data Alvo"].dt.to_period("M").astype(str).replace("NaT", "Sem data")
    df["Semana Alvo"] = df["Data Alvo"].dt.strftime("%d/%m/%Y").fillna("Sem data")
    df["Inicio Gantt"] = df["Data Alvo"] - pd.Timedelta(days=7)
    df["Fim Gantt"] = df["Data Entrega"].fillna(df["Data Alvo"])
    df.loc[df["Fim Gantt"] < df["Inicio Gantt"], "Fim Gantt"] = df["Inicio Gantt"] + pd.Timedelta(days=1)
    return df


def pct(parte, total):
    return (parte / total * 100) if total else 0


def card(titulo, valor, subtitulo, cor="#0F172A", icone=""):
    st.markdown(
        f"""
        <div class="card">
            <div class="metric-label">{icone} {titulo}</div>
            <div class="metric-value">{valor}</div>
            <div class="metric-delta" style="color:{cor};">{subtitulo}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def gerar_excel(df, resumo):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Atividades")
        resumo.to_excel(writer, index=False, sheet_name="Resumo")
    return output.getvalue()


def limpar_texto(valor, padrao="Não informado"):
    """Converte valores vindos do Excel para texto seguro, evitando mistura float/str nos filtros."""
    if pd.isna(valor):
        return padrao
    texto = str(valor).strip()
    if texto == "" or texto.lower() in ["nan", "nat", "none"]:
        return padrao
    # remove .0 de números inteiros vindos do Excel, ex.: 10.0 -> 10
    try:
        numero = float(texto.replace(",", "."))
        if numero.is_integer():
            return str(int(numero))
    except Exception:
        pass
    return texto


def lista_filtro(df, coluna):
    """Lista segura para multiselect: remove nulos, converte tudo para string e ordena."""
    if coluna not in df.columns:
        return []
    valores = [limpar_texto(v) for v in df[coluna].tolist()]
    return sorted(set(v for v in valores if v))


def arquivo_upload_para_base64(arquivo):
    """Converte imagem carregada no Streamlit para base64, permitindo posicionamento no cabeçalho HTML."""
    if arquivo is None:
        return None
    try:
        return base64.b64encode(arquivo.getvalue()).decode("utf-8")
    except Exception:
        return None



def nome_arquivo_seguro(nome: str) -> str:
    """Converte o nome do projeto em um nome de arquivo seguro."""
    nome = limpar_texto(nome, padrao="Cronograma")
    nome = re.sub(r"[^A-Za-zÀ-ÿ0-9 _.-]+", "", nome).strip()
    nome = re.sub(r"\s+", "_", nome)
    return nome[:80] or "Cronograma"


def caminho_cronograma_por_nome(nome_projeto: str) -> Path:
    """Retorna o caminho .xlsx para salvar/carregar um cronograma criado no sistema."""
    return CRONOGRAMAS_DIR / f"{nome_arquivo_seguro(nome_projeto)}{EXTENSAO_CRONOGRAMA_SISTEMA}"


def listar_cronogramas_salvos() -> list[Path]:
    """Lista cronogramas já salvos localmente pelo sistema."""
    if not CRONOGRAMAS_DIR.exists():
        return []
    return sorted(CRONOGRAMAS_DIR.glob(f"*{EXTENSAO_CRONOGRAMA_SISTEMA}"), key=lambda p: p.stat().st_mtime, reverse=True)


def ler_cronograma_criado_no_sistema(arquivo_origem) -> pd.DataFrame:
    """Lê cronograma salvo/criado pelo sistema, aceitando arquivo local ou upload."""
    try:
        if hasattr(arquivo_origem, "read"):
            dados = arquivo_origem.getvalue()
            xls = pd.ExcelFile(BytesIO(dados))
        else:
            xls = pd.ExcelFile(Path(arquivo_origem))

        # Preferência para a aba padrão dos cronogramas criados no sistema.
        aba = "Cronograma" if "Cronograma" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=aba)

        # Também aceita arquivos exportados anteriormente com aba Atividades.
        if "Atividades" in xls.sheet_names and "Atividade" not in df.columns:
            df = pd.read_excel(xls, sheet_name="Atividades")

        return df
    except Exception as erro:
        st.error(f"Erro ao abrir cronograma criado no sistema: {erro}")
        st.stop()


def salvar_cronograma_excel_sistema(df_cronograma: pd.DataFrame, nome_projeto: str) -> Path:
    """Salva o cronograma editável no padrão que o próprio sistema consegue reabrir."""
    CRONOGRAMAS_DIR.mkdir(exist_ok=True)
    caminho = caminho_cronograma_por_nome(nome_projeto)
    df_salvar = df_cronograma.copy()
    df_salvar.insert(0, "Projeto", limpar_texto(nome_projeto, padrao="Cronograma"))

    with pd.ExcelWriter(caminho, engine="openpyxl", date_format="DD/MM/YYYY", datetime_format="DD/MM/YYYY") as writer:
        df_salvar.to_excel(writer, index=False, sheet_name="Cronograma")
        preparar_dataframe_cronograma(df_cronograma).to_excel(writer, index=False, sheet_name="Atividades")

    # Mantém compatibilidade com a versão anterior, que carregava cronograma_ativo.xlsx.
    shutil.copy2(caminho, CRONOGRAMA_ATIVO)
    return caminho

def preparar_dataframe_cronograma(df_entrada: pd.DataFrame) -> pd.DataFrame:
    """Padroniza dados criados no sistema para usar nos mesmos indicadores do dashboard."""
    df = df_entrada.copy()

    # Correção: versões anteriores usavam o nome "Data Atividade".
    # O restante do dashboard usa "Data Entrega", então padronizamos aqui.
    if "Data Entrega" not in df.columns and "Data Atividade" in df.columns:
        df = df.rename(columns={"Data Atividade": "Data Entrega"})

    colunas = ["Item", "Fase", "Processo", "Atividade", "Responsável", "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Observações"]

    for col in colunas:
        if col not in df.columns:
            df[col] = pd.NaT if col in ["Data Início", "Data Alvo", "Data Entrega"] else ""

    df = df[colunas].copy()
    df["Atividade"] = df["Atividade"].apply(lambda x: limpar_texto(x, padrao=""))
    df = df[df["Atividade"].astype(str).str.strip() != ""].copy()

    if df.empty:
        return pd.DataFrame(columns=[
            "Linha", "Item", "Fase", "Processo", "Atividade", "Responsável",
            "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Observações", "Status", "Dias para Vencer",
            "Mês Alvo", "Semana Alvo", "Inicio Gantt", "Fim Gantt"
        ])

    for col in ["Item", "Fase", "Processo", "Responsável", "E-mail Responsável", COL_DEPENDENCIA, "Observações"]:
        padrao = "Não informado" if col in ["Fase", "Responsável"] else ""
        df[col] = df[col].apply(lambda x: limpar_texto(x, padrao=padrao))

    df["Linha"] = range(1, len(df) + 1)
    df["Data Início"] = pd.to_datetime(df["Data Início"], errors="coerce", dayfirst=True)
    df["Data Alvo"] = pd.to_datetime(df["Data Alvo"], errors="coerce", dayfirst=True)
    df["Data Entrega"] = pd.to_datetime(df["Data Entrega"], errors="coerce", dayfirst=True)
    df["Status"] = df.apply(calcular_status, axis=1)

    hoje = pd.Timestamp.now().normalize()
    df["Dias para Vencer"] = pd.NA
    mask_data_alvo = df["Data Alvo"].notna()
    df.loc[mask_data_alvo, "Dias para Vencer"] = (
        df.loc[mask_data_alvo, "Data Alvo"] - hoje
    ).dt.days.astype("Int64")

    df["Mês Alvo"] = df["Data Alvo"].dt.to_period("M").astype(str).replace("NaT", "Sem data")
    df["Semana Alvo"] = df["Data Alvo"].dt.strftime("%d/%m/%Y").fillna("Sem data")
    df["Inicio Gantt"] = df["Data Alvo"] - pd.Timedelta(days=7)
    df["Fim Gantt"] = df["Data Entrega"].fillna(df["Data Alvo"])
    df.loc[df["Fim Gantt"] < df["Inicio Gantt"], "Fim Gantt"] = df["Inicio Gantt"] + pd.Timedelta(days=1)
    return df



def normalizar_tipos_editor_cronograma(df_entrada: pd.DataFrame) -> pd.DataFrame:
    """Garante tipos compatíveis com st.data_editor.

    O Streamlit não permite configurar TextColumn sobre coluna que o pandas
    identificou como INTEGER. Por isso, campos editáveis de texto são
    convertidos explicitamente para string antes de abrir o editor.
    """
    df = df_entrada.copy() if df_entrada is not None else pd.DataFrame()

    colunas_texto = ["Item", "Fase", "Processo", "Atividade", "Responsável", "E-mail Responsável", COL_DEPENDENCIA, "Observações"]
    colunas_data = ["Data Início", "Data Alvo", "Data Entrega"]
    colunas_base = ["Item", "Fase", "Processo", "Atividade", "Responsável", "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Observações"]

    for col in colunas_base:
        if col not in df.columns:
            df[col] = pd.NaT if col in colunas_data else ""

    # Remove coluna Projeto do editor; ela é controlada pelo campo Nome do cronograma / Projeto.
    if "Projeto" in df.columns:
        df = df.drop(columns=["Projeto"])

    # Compatibilidade com nomes anteriores/alternativos da coluna de dependência.
    aliases_dependencia = [
        "Dependência",
        "Dependencia",
        "Depende da Entrega",
        "Depende da Entrega da Atividade",
        "Depende da Atividade",
    ]
    if COL_DEPENDENCIA not in df.columns:
        for alias in aliases_dependencia:
            if alias in df.columns:
                df[COL_DEPENDENCIA] = df[alias]
                break

    for col in colunas_texto:
        df[col] = df[col].apply(lambda x: "" if pd.isna(x) else str(x).strip())
        # Evita exibição de 1.0, 2.0 etc quando o Excel devolve número decimal.
        df[col] = df[col].str.replace(r"^(-?\d+)\.0$", r"\1", regex=True)

    for col in colunas_data:
        df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

    return df[colunas_base].copy()

def modelo_cronograma_vazio(nome_projeto: str = "") -> pd.DataFrame:
    """Modelo inicial para criação de cronograma na interface."""
    projeto = limpar_texto(nome_projeto, padrao="")
    return normalizar_tipos_editor_cronograma(pd.DataFrame({
        "Item": ["1", "2", "3"],
        "Fase": ["0", "1", "2"],
        "Processo": ["Planejamento", "Execução", "Validação"],
        "Atividade": ["Definir escopo do projeto", "Executar atividades principais", "Validar entrega com cliente"],
        "Responsável": ["Não informado", "Não informado", "Não informado"],
        "E-mail Responsável": ["", "", ""],
        "Data Início": [pd.NaT, pd.NaT, pd.NaT],
        "Data Alvo": [pd.NaT, pd.NaT, pd.NaT],
        "Data Entrega": [pd.NaT, pd.NaT, pd.NaT],
        COL_DEPENDENCIA: ["", "", ""],
        "Observações": ["", "", ""],
    }))


def carregar_cronograma_local(nome_projeto: str = "") -> pd.DataFrame:
    """Carrega cronograma salvo pelo nome informado; se não existir, abre o último ativo ou modelo vazio."""
    caminho_nomeado = caminho_cronograma_por_nome(nome_projeto) if nome_projeto else None

    for caminho in [caminho_nomeado, CRONOGRAMA_ATIVO]:
        if caminho and caminho.exists():
            try:
                df = ler_cronograma_criado_no_sistema(caminho)
                return normalizar_tipos_editor_cronograma(df)
            except Exception:
                st.warning("Não foi possível carregar o cronograma salvo. Um modelo vazio será aberto.")
                break

    return modelo_cronograma_vazio(nome_projeto)


def salvar_cronograma_local(df_cronograma: pd.DataFrame, nome_projeto: str) -> Path:
    """Salva o cronograma criado na interface em arquivo Excel local, com o nome do projeto."""
    return salvar_cronograma_excel_sistema(df_cronograma, nome_projeto)


def editor_cronograma_interface(nome_projeto: str = "") -> tuple[pd.DataFrame, str]:
    """Exibe uma planilha editável no Streamlit e retorna dados tratados e nome do projeto."""
    st.markdown("<div class='section-title'>Criar / Editar Cronograma do Projeto</div>", unsafe_allow_html=True)
    st.caption("Informe o nome do cronograma/projeto e preencha as atividades diretamente na tela. O status será calculado automaticamente pelas datas alvo e de conclusão.")

    nome_projeto = limpar_texto(nome_projeto, padrao="Cronograma")
    nome_projeto = st.text_input(
        "Nome do cronograma / Projeto",
        value=st.session_state.get("nome_projeto_cronograma", nome_projeto),
        key="input_nome_projeto_cronograma",
        help="Este nome será exibido no campo Projeto e usado como nome do arquivo ao salvar.",
    ).strip() or "Cronograma"
    st.session_state["nome_projeto_cronograma"] = nome_projeto

    chave_editor = f"cronograma_editor_{nome_arquivo_seguro(nome_projeto)}"
    if chave_editor not in st.session_state:
        st.session_state[chave_editor] = carregar_cronograma_local(nome_projeto)

    # Correção: força tipos compatíveis com column_config do st.data_editor.
    st.session_state[chave_editor] = normalizar_tipos_editor_cronograma(st.session_state[chave_editor])

    itens_dependencia = [
        limpar_texto(v, padrao="")
        for v in st.session_state[chave_editor].get("Item", pd.Series(dtype=str)).tolist()
    ]
    opcoes_dependencia = [""] + sorted(set(v for v in itens_dependencia if v))

    df_editado = st.data_editor(
        st.session_state[chave_editor],
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
        column_config={
            "Item": st.column_config.TextColumn("Item"),
            "Fase": st.column_config.SelectboxColumn("Fase", options=["0", "1", "2", "3", "4", "5"], required=True),
            "Processo": st.column_config.TextColumn("Processo"),
            "Atividade": st.column_config.TextColumn("Atividade", required=True),
            "Responsável": st.column_config.TextColumn("Responsável"),
            "E-mail Responsável": st.column_config.TextColumn("E-mail responsável"),
            "Data Início": st.column_config.DateColumn("Data Início", format="DD/MM/YYYY"),
            "Data Alvo": st.column_config.DateColumn("Data Alvo", format="DD/MM/YYYY"),
            "Data Entrega": st.column_config.DateColumn("Data Entrega", format="DD/MM/YYYY"),
            COL_DEPENDENCIA: st.column_config.SelectboxColumn(
                "Depende da Entrega da Atividade (Item)",
                options=opcoes_dependencia,
                help="Selecione o Item da atividade que precisa ser entregue antes desta. Deixe em branco quando não houver dependência.",
            ),
            "Observações": st.column_config.TextColumn("Observações / anexos"),
        },
        key=f"data_editor_cronograma_{nome_arquivo_seguro(nome_projeto)}",
    )

    # Anexos de entrega: o st.data_editor ainda não possui coluna nativa para upload de arquivo.
    # Por isso, os documentos são anexados por atividade e o nome dos arquivos fica registrado em Observações.
    with st.expander("📎 Anexar documentos referentes à entrega"):
        if df_editado.empty:
            st.info("Inclua uma atividade antes de anexar documentos.")
        else:
            opcoes_anexo = [
                f"{idx} | Item {limpar_texto(row.get('Item', ''), padrao='')} | {limpar_texto(row.get('Atividade', ''), padrao='Atividade')[:80]}"
                for idx, row in df_editado.iterrows()
            ]
            atividade_anexo = st.selectbox(
                "Selecionar atividade para vincular documentos",
                opcoes_anexo,
                key=f"atividade_anexo_{nome_arquivo_seguro(nome_projeto)}",
            )
            arquivos_anexo = st.file_uploader(
                "Anexar documento(s) da entrega",
                type=["pdf", "doc", "docx", "xls", "xlsx", "png", "jpg", "jpeg", "txt", "zip"],
                accept_multiple_files=True,
                key=f"upload_docs_entrega_{nome_arquivo_seguro(nome_projeto)}",
            )
            if st.button("📎 Salvar anexos na atividade", width="stretch", key=f"btn_salvar_anexos_{nome_arquivo_seguro(nome_projeto)}"):
                if not arquivos_anexo:
                    st.warning("Selecione pelo menos um documento para anexar.")
                else:
                    idx_anexo = int(str(atividade_anexo).split(" | ")[0])
                    pasta_docs = DOCUMENTOS_ENTREGA_DIR / nome_arquivo_seguro(nome_projeto)
                    pasta_docs.mkdir(parents=True, exist_ok=True)
                    nomes_salvos = []
                    for arq in arquivos_anexo:
                        nome_doc = nome_arquivo_seguro(Path(arq.name).stem) + Path(arq.name).suffix.lower()
                        destino = pasta_docs / nome_doc
                        contador = 1
                        while destino.exists():
                            destino = pasta_docs / f"{Path(nome_doc).stem}_{contador}{Path(nome_doc).suffix}"
                            contador += 1
                        destino.write_bytes(arq.getvalue())
                        nomes_salvos.append(destino.name)

                    obs_atual = limpar_texto(df_editado.at[idx_anexo, "Observações"], padrao="")
                    texto_anexos = "Anexos: " + ", ".join(nomes_salvos)
                    df_editado.at[idx_anexo, "Observações"] = (obs_atual + "\n" if obs_atual else "") + texto_anexos
                    st.session_state[chave_editor] = normalizar_tipos_editor_cronograma(df_editado)
                    st.success("Documento(s) anexado(s) e registrados em Observações.")
                    st.rerun()

    c_salvar, c_limpar, c_download = st.columns([1, 1, 1])
    with c_salvar:
        if st.button("💾 Salvar cronograma", width="stretch"):
            st.session_state[chave_editor] = df_editado
            caminho = salvar_cronograma_local(df_editado, nome_projeto)
            st.success(f"Cronograma salvo com sucesso: {caminho.name}")
            st.rerun()
    with c_limpar:
        if st.button("🧹 Limpar modelo", width="stretch"):
            st.session_state[chave_editor] = normalizar_tipos_editor_cronograma(pd.DataFrame({
                "Item": [""],
                "Fase": ["0"],
                "Processo": [""],
                "Atividade": [""],
                "Responsável": [""],
                "E-mail Responsável": [""],
                "Data Início": [pd.NaT],
                "Data Alvo": [pd.NaT],
                "Data Entrega": [pd.NaT],
                COL_DEPENDENCIA: [""],
                "Observações": [""],
            }))
            st.rerun()
    with c_download:
        df_download = df_editado.copy()
        df_download.insert(0, "Projeto", nome_projeto)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl", date_format="DD/MM/YYYY", datetime_format="DD/MM/YYYY") as writer:
            df_download.to_excel(writer, index=False, sheet_name="Cronograma")
            preparar_dataframe_cronograma(df_editado).to_excel(writer, index=False, sheet_name="Atividades")
        st.download_button(
            "⬇️ Baixar cronograma",
            data=output.getvalue(),
            file_name=f"{nome_arquivo_seguro(nome_projeto)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    return preparar_dataframe_cronograma(df_editado), nome_projeto



def email_valido(email: str) -> bool:
    """Valida formato básico de e-mail para alertas."""
    email = limpar_texto(email, padrao="")
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def montar_email_alerta(row, nome_projeto: str, cliente: str = "") -> tuple[str, str]:
    """Monta assunto e corpo do alerta de atividade."""
    atividade = limpar_texto(row.get("Atividade", ""), padrao="Atividade não informada")
    status = limpar_texto(row.get("Status", ""), padrao="Status não informado")
    responsavel = limpar_texto(row.get("Responsável", ""), padrao="Responsável não informado")
    data_alvo = row.get("Data Alvo")
    data_entrega = row.get("Data Entrega")
    dias = row.get("Dias para Vencer", "")

    data_alvo_txt = data_alvo.strftime("%d/%m/%Y") if pd.notna(data_alvo) else "Sem data"
    data_entrega_txt = data_entrega.strftime("%d/%m/%Y") if pd.notna(data_entrega) else "Não concluída"

    assunto = f"[PMO] Alerta de atividade - {status} - {atividade[:60]}"
    corpo = f"""Olá {responsavel},

Este é um alerta automático do Dashboard PMO referente à atividade sob sua responsabilidade.

Projeto: {nome_projeto}
Cliente: {cliente if cliente else "Cliente não informado"}
Atividade: {atividade}
Responsável: {responsavel}
Status atual: {status}
Data alvo: {data_alvo_txt}
Data entrega/conclusão: {data_entrega_txt}
Dias para vencer/atraso: {dias if str(dias) not in ["<NA>", "nan", "NaT"] else "Não calculado"}

Favor avaliar a situação da atividade e atualizar o cronograma quando houver nova informação.

Atenciosamente,
Dashboard PMO - AGSolution
"""
    return assunto, corpo


def link_mailto(destinatario: str, assunto: str, corpo: str) -> str:
    """Gera link mailto para abrir Outlook/Gmail/app de e-mail padrão."""
    return (
        "mailto:"
        + urllib.parse.quote(destinatario)
        + "?subject="
        + urllib.parse.quote(assunto)
        + "&body="
        + urllib.parse.quote(corpo)
    )


def pagina_alertas_email(df_origem: pd.DataFrame, nome_projeto: str, cliente: str = "") -> None:
    """Página para preparar alertas de e-mail por responsável."""
    st.markdown("<div class='section-title'>Alertas por E-mail</div>", unsafe_allow_html=True)
    st.caption("Use esta tela para localizar atividades por status e abrir o e-mail padrão já preenchido para o responsável.")

    if "E-mail Responsável" not in df_origem.columns:
        st.warning("O cronograma ainda não possui a coluna E-mail Responsável.")
        return

    status_alerta = st.multiselect(
        "Status para alertar",
        options=["Em Atraso", "Em Andamento", "Concluida atrasada", "Sem data alvo", "Concluida"],
        default=["Em Atraso", "Concluida atrasada", "Em Andamento"],
    )

    somente_com_email = st.checkbox("Mostrar somente atividades com e-mail válido", value=True)

    alertas = df_origem[df_origem["Status"].isin(status_alerta)].copy()
    alertas["E-mail válido"] = alertas["E-mail Responsável"].apply(email_valido)

    if somente_com_email:
        alertas = alertas[alertas["E-mail válido"]].copy()

    if alertas.empty:
        st.info("Nenhuma atividade encontrada para os critérios selecionados.")
        return

    colunas_alerta = [
        "Item", "Fase", "Processo", "Atividade", "Responsável",
        "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Dias para Vencer", "Status", "E-mail válido"
    ]
    st.dataframe(alertas[colunas_alerta], width="stretch", hide_index=True)

    st.divider()
    st.subheader("Gerar e-mail de alerta")

    opcoes = [
        f"{idx} | {limpar_texto(row.get('Status'))} | {limpar_texto(row.get('Responsável'))} | {limpar_texto(row.get('Atividade'))[:70]}"
        for idx, row in alertas.iterrows()
    ]

    escolha = st.selectbox("Selecione a atividade", opcoes)
    idx_sel = int(str(escolha).split(" | ")[0])
    row = alertas.loc[idx_sel]
    destinatario = limpar_texto(row.get("E-mail Responsável", ""), padrao="")

    assunto, corpo = montar_email_alerta(row, nome_projeto, cliente)
    st.text_input("Para", value=destinatario, disabled=True)
    st.text_input("Assunto", value=assunto, disabled=True)
    st.text_area("Mensagem", value=corpo, height=280)

    if email_valido(destinatario):
        url = link_mailto(destinatario, assunto, corpo)
        st.markdown(
            f"<a href='{url}' target='_blank'><button style='background:#2563EB;color:white;border:none;border-radius:10px;padding:10px 16px;font-weight:700;'>📧 Abrir e-mail para envio</button></a>",
            unsafe_allow_html=True,
        )
        st.info("O botão abre o Outlook/Gmail/app de e-mail padrão com a mensagem preenchida. O envio final fica sob confirmação do usuário.")
    else:
        st.warning("Esta atividade não possui e-mail válido do responsável.")


# =========================================================
# SIDEBAR
# =========================================================
with st.sidebar:
    st.markdown("<div class='sidebar-logo-box'>", unsafe_allow_html=True)
    if LOGO_SIDEBAR.exists():
        st.image(str(LOGO_SIDEBAR), width=230)
    else:
        st.markdown("## AGSolution")
        st.caption("Arquivo AGSolution_logo.png não encontrado na pasta do programa.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(
        """
        <div class='sidebar-caption'>
            
        </div>
        """,
        unsafe_allow_html=True
    )

    st.divider()

    opcoes_menu = ["Visão Geral", "Cronograma", "Projetos", "Indicadores", "Riscos", "Alertas E-mail", "Relatórios", "Exportar", "Configurações"]
    if usuario_eh_admin():
        opcoes_menu.append("Administração de Usuários")

    menu = st.radio(
        "Menu",
        opcoes_menu,
        index=0
    )

    st.caption(f"Logado como: {st.session_state.get('nome_logado', '')} | {st.session_state.get('perfil_logado', '')}")
    if st.button("Sair", width="stretch"):
        fazer_logout()

    st.divider()
    modo_dados = st.radio(
        "Fonte dos dados",
        [
            "Criar/editar cronograma no sistema",
            "Upload cronograma criado no sistema",
        ],
        index=0,
        help="A aplicação agora aceita somente cronogramas criados no próprio sistema ou upload de cronograma exportado pelo sistema.",
    )

    arquivo_cronograma_sistema = None
    nome_projeto_sidebar = ""
    continuar_editando_upload = True

    if modo_dados == "Criar/editar cronograma no sistema":
        nome_projeto_sidebar = st.text_input(
            "Nome do cronograma / Projeto",
            value=st.session_state.get("nome_projeto_cronograma", "Novo Projeto"),
            key="nome_projeto_sidebar",
        )
        cronogramas_salvos = listar_cronogramas_salvos()
        if cronogramas_salvos:
            nomes_salvos = [p.stem for p in cronogramas_salvos if p.name != CRONOGRAMA_ATIVO.name]
            if nomes_salvos:
                selecionado = st.selectbox("Abrir cronograma salvo", ["Novo / manter nome digitado"] + nomes_salvos)
                if selecionado != "Novo / manter nome digitado":
                    nome_projeto_sidebar = selecionado
                    st.session_state["nome_projeto_cronograma"] = selecionado
        st.info("O cronograma será criado, editado e salvo pelo nome do projeto.")
    else:
        arquivo_cronograma_sistema = st.file_uploader(
            "Upload cronograma criado no sistema",
            type=["xlsx"],
            help="Use esta opção para abrir arquivos baixados/salvos pela opção Criar/editar cronograma no sistema.",
        )
        continuar_editando_upload = st.checkbox(
            "Continuar editando após upload",
            value=True,
            help="Quando marcado, o arquivo carregado será aberto no editor para alterar e salvar novamente.",
        )

    st.divider()
    cliente_nome = st.text_input(
        "Nome do Cliente",
        value="",
        key="cliente_nome"
    )

    logo_cliente = st.file_uploader(
        "Carregar logo do cliente",
        type=["png", "jpg", "jpeg"],
        key="logo_cliente"
    )

# =========================================================
# CARREGAMENTO
# =========================================================
if modo_dados == "Criar/editar cronograma no sistema":
    df_base, nome_arquivo_exibicao = editor_cronograma_interface(nome_projeto_sidebar)
    if df_base.empty:
        st.warning("Inclua pelo menos uma atividade para iniciar os indicadores do Dashboard PMO.")
        st.stop()
else:
    if arquivo_cronograma_sistema is None:
        st.warning("Carregue um cronograma criado pelo sistema para iniciar o Dashboard PMO.")
        st.stop()

    df_cronograma_upload = ler_cronograma_criado_no_sistema(arquivo_cronograma_sistema)

    if "Projeto" in df_cronograma_upload.columns and df_cronograma_upload["Projeto"].notna().any():
        nome_arquivo_exibicao = limpar_texto(
            df_cronograma_upload["Projeto"].dropna().iloc[0],
            padrao=Path(arquivo_cronograma_sistema.name).stem,
        )
        df_cronograma_upload = df_cronograma_upload.drop(columns=["Projeto"])
    else:
        nome_arquivo_exibicao = Path(arquivo_cronograma_sistema.name).stem

    # Nova opção: abrir o cronograma enviado diretamente no editor,
    # permitindo continuar editando e salvar novamente pelo nome do projeto.
    if continuar_editando_upload:
        nome_projeto_upload = nome_arquivo_exibicao or Path(arquivo_cronograma_sistema.name).stem or "Cronograma"
        st.session_state["nome_projeto_cronograma"] = nome_projeto_upload

        chave_editor_upload = f"cronograma_editor_{nome_arquivo_seguro(nome_projeto_upload)}"
        st.session_state[chave_editor_upload] = normalizar_tipos_editor_cronograma(df_cronograma_upload)

        df_base, nome_arquivo_exibicao = editor_cronograma_interface(nome_projeto_upload)
    else:
        df_base = preparar_dataframe_cronograma(df_cronograma_upload)

    if df_base.empty:
        st.warning("O cronograma carregado não possui atividades válidas.")
        st.stop()

# =========================================================
# CABEÇALHO EXECUTIVO E FILTROS GLOBAIS
# =========================================================
ultima_atualizacao = datetime.now().strftime("%d/%m/%Y %H:%M")
cliente_nome_seguro = html.escape(cliente_nome.strip() if cliente_nome else "Cliente não informado")
nome_arquivo_seguro = html.escape(nome_arquivo_exibicao)
logo_cliente_b64 = arquivo_upload_para_base64(logo_cliente)

if logo_cliente_b64:
    logo_cliente_html = f"<img src='data:image/png;base64,{logo_cliente_b64}' alt='Logo do cliente'>"
else:
    logo_cliente_html = "<div class='executive-logo-placeholder'>Logo do cliente</div>"

st.markdown(
    f"""
    <div class='executive-header'>
        <div class='executive-logo-area'>
            {logo_cliente_html}
        </div>
        <div class='executive-main-area'>
            <div class='executive-title-row'>
                <div>
                    <div class='executive-title'>Dashboard PMO</div>
                    <div class='executive-subtitle'>Visão geral do desempenho dos projetos e cronograma</div>
                </div>
                <div class='executive-version'>{VERSAO_DASHBOARD}</div>
            </div>
            <div class='executive-meta'>
                <div><b>Cliente:</b> {cliente_nome_seguro}</div>
                <div><b>Última atualização:</b> {ultima_atualizacao}</div>
                <div style='grid-column: 1 / -1;'><b>Projeto:</b> {nome_arquivo_seguro}</div>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

f1, f2, f3, f4 = st.columns([1.2, 1.2, 1.2, 1.8])
with f1:
    status_opcoes = lista_filtro(df_base, "Status")
    status_sel = st.multiselect("Status", options=status_opcoes, default=status_opcoes)
with f2:
    resp_opcoes = lista_filtro(df_base, "Responsável")
    resp_sel = st.multiselect("Responsável", options=resp_opcoes, default=resp_opcoes)
with f3:
    fase_opcoes = lista_filtro(df_base, "Fase")
    fase_sel = st.multiselect("Fase", options=fase_opcoes, default=fase_opcoes)
with f4:
    datas_validas = df_base["Data Alvo"].dropna()
    if not datas_validas.empty:
        periodo = st.date_input(
            "Período Data Alvo",
            value=(datas_validas.min().date(), datas_validas.max().date()),
            format="DD/MM/YYYY"
        )
    else:
        periodo = None

df = df_base[
    df_base["Status"].isin(status_sel)
    & df_base["Responsável"].isin(resp_sel)
    & df_base["Fase"].isin(fase_sel)
].copy()

if periodo and len(periodo) == 2:
    inicio, fim = pd.Timestamp(periodo[0]), pd.Timestamp(periodo[1])
    df = df[(df["Data Alvo"].isna()) | ((df["Data Alvo"] >= inicio) & (df["Data Alvo"] <= fim))]

# =========================================================
# MÉTRICAS
# =========================================================
total = len(df)
qtd_concluida = int((df["Status"] == "Concluida").sum())
qtd_andamento = int((df["Status"] == "Em Andamento").sum())
qtd_atraso = int((df["Status"] == "Em Atraso").sum())
qtd_concluida_atrasada = int((df["Status"] == "Concluida atrasada").sum())
qtd_atencao = qtd_andamento + qtd_concluida_atrasada
conclusao_geral = pct(qtd_concluida + qtd_concluida_atrasada, total)

resumo = df.groupby("Status", dropna=False).size().reset_index(name="Quantidade")
resumo["Percentual"] = (resumo["Quantidade"] / total * 100).round(1) if total else 0
resumo["Status Curto"] = resumo["Status"].map(STATUS_CURTO).fillna(resumo["Status"])

m1, m2, m3, m4, m5 = st.columns(5)
with m1: card("Total de Atividades", total, "100% do total", "#2563EB", "📋")
with m2: card("Atrasadas", qtd_atraso, f"{pct(qtd_atraso, total):.1f}% do total", "#DC2626", "⏰")
with m3: card("Em Atenção", qtd_atencao, f"{pct(qtd_atencao, total):.1f}% do total", "#D97706", "⚠️")
with m4: card("No Prazo", qtd_concluida, f"{pct(qtd_concluida, total):.1f}% do total", "#16A34A", "✅")
with m5: card("Conclusão Geral", f"{conclusao_geral:.0f}%", "% de atividades concluídas", "#2563EB", "🎯")

# =========================================================
# PÁGINAS
# =========================================================
if menu == "Visão Geral":
    c1, c2 = st.columns([1, 1.35])
    with c1:
        st.markdown("<div class='section-title'>Status das Atividades</div>", unsafe_allow_html=True)
        fig_pizza = px.pie(
            resumo,
            values="Quantidade",
            names="Status Curto",
            hole=0.52,
            color="Status",
            color_discrete_map=CORES_STATUS,
        )
        fig_pizza.update_traces(textinfo="percent", textfont_size=15)
        fig_pizza.update_layout(height=390, margin=dict(l=10, r=10, t=20, b=10), showlegend=True)
        st.plotly_chart(fig_pizza, width="stretch")

    with c2:
        st.markdown("<div class='section-title'>Evolução de Conclusão (%)</div>", unsafe_allow_html=True)
        evolucao = df.dropna(subset=["Data Alvo"]).sort_values("Data Alvo").copy()
        evolucao["Concluida"] = evolucao["Status"].isin(["Concluida", "Concluida atrasada"]).astype(int)
        if not evolucao.empty:
            evolucao["Acumulado"] = evolucao["Concluida"].cumsum() / pd.Series(range(1, len(evolucao) + 1), index=evolucao.index)
            evolucao["Conclusão %"] = (evolucao["Acumulado"] * 100).round(1)
            fig_linha = px.line(evolucao, x="Data Alvo", y="Conclusão %", markers=True)
            fig_linha.update_layout(height=390, yaxis_range=[0, 100], margin=dict(l=10, r=10, t=20, b=10))
            st.plotly_chart(fig_linha, width="stretch")
        else:
            st.info("Sem datas alvo para gerar evolução.")

    c3, c4, c5 = st.columns([1.45, 0.7, 0.85])
    with c3:
        st.markdown("<div class='section-title'>Cronograma Macro (Gantt)</div>", unsafe_allow_html=True)
        gantt = df.dropna(subset=["Inicio Gantt", "Fim Gantt"]).head(35)
        if not gantt.empty:
            fig_gantt = px.timeline(
                gantt,
                x_start="Inicio Gantt",
                x_end="Fim Gantt",
                y="Atividade",
                color="Status",
                color_discrete_map=CORES_STATUS,
                hover_data=["Item", "Fase", "Processo", "Responsável", "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA],
            )
            fig_gantt.update_yaxes(autorange="reversed")
            fig_gantt.update_layout(height=380, margin=dict(l=10, r=10, t=20, b=10), showlegend=False)
            st.plotly_chart(fig_gantt, width="stretch")
        else:
            st.info("Sem dados para Gantt.")

    with c4:
        st.markdown("<div class='section-title'>Atividades por Responsável</div>", unsafe_allow_html=True)
        por_resp = df.groupby("Responsável").size().reset_index(name="Quantidade").sort_values("Quantidade")
        fig_resp = px.bar(por_resp, x="Quantidade", y="Responsável", orientation="h", text="Quantidade")
        fig_resp.update_layout(height=380, margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig_resp, width="stretch")

    with c5:
        st.markdown("<div class='section-title'>Principais Riscos</div>", unsafe_allow_html=True)
        riscos = df[df["Status"].isin(["Em Atraso", "Concluida atrasada"])]\
            .sort_values(["Status", "Dias para Vencer"]).head(5)
        if riscos.empty:
            st.success("Nenhum risco crítico no filtro atual.")
        else:
            for _, r in riscos.iterrows():
                st.markdown(
                    f"**🔴 {r['Atividade']}**  \n"
                    f"Status: {r['Status']}  \n"
                    f"Responsável: {r['Responsável']}  \n"
                    f"Alvo: {r['Data Alvo'].strftime('%d/%m/%Y') if pd.notna(r['Data Alvo']) else 'Sem data'}"
                )
                st.divider()

elif menu == "Cronograma":
    st.markdown("<div class='section-title'>Cronograma Detalhado</div>", unsafe_allow_html=True)
    gantt = df.dropna(subset=["Inicio Gantt", "Fim Gantt"])
    if gantt.empty:
        st.info("Sem dados para exibir no Gantt com o filtro atual.")
    else:
        fig_gantt = px.timeline(
            gantt,
            x_start="Inicio Gantt",
            x_end="Fim Gantt",
            y="Atividade",
            color="Status",
            color_discrete_map=CORES_STATUS,
            hover_data=["Item", "Fase", "Processo", "Responsável", "E-mail Responsável", "Data Início", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Dias para Vencer"],
        )
        fig_gantt.update_yaxes(autorange="reversed")
        fig_gantt.update_layout(height=max(500, len(gantt) * 18), margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig_gantt, width="stretch")

elif menu == "Projetos":
    st.markdown("<div class='section-title'>Visão por Fase / Projeto</div>", unsafe_allow_html=True)
    matriz = pd.crosstab(df["Fase"], df["Status"]).reset_index()
    st.dataframe(matriz, width="stretch", hide_index=True)
    fig_fase = px.bar(df, x="Fase", color="Status", color_discrete_map=CORES_STATUS, barmode="stack")
    st.plotly_chart(fig_fase, width="stretch")

elif menu == "Indicadores":
    st.markdown("<div class='section-title'>Indicadores Avançados</div>", unsafe_allow_html=True)
    ind1, ind2 = st.columns(2)
    with ind1:
        fig_mes = px.histogram(df.dropna(subset=["Data Alvo"]), x="Mês Alvo", color="Status", color_discrete_map=CORES_STATUS, barmode="group")
        fig_mes.update_layout(title="Atividades por Mês Alvo")
        st.plotly_chart(fig_mes, width="stretch")
    with ind2:
        heat = pd.crosstab(df["Responsável"], df["Status"])
        fig_heat = px.imshow(heat, text_auto=True, aspect="auto", title="Mapa de Calor por Responsável x Status")
        st.plotly_chart(fig_heat, width="stretch")

elif menu == "Riscos":
    st.markdown("<div class='section-title'>Riscos e Pendências</div>", unsafe_allow_html=True)
    riscos = df[df["Status"].isin(["Em Atraso", "Concluida atrasada", "Sem data alvo"])].copy()
    riscos["Criticidade"] = riscos["Status"].map({"Em Atraso": "Alto", "Concluida atrasada": "Médio", "Sem data alvo": "Médio"}).fillna("Baixo")
    st.dataframe(riscos[["Item", "Atividade", "Responsável", "E-mail Responsável", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Status", "Criticidade"]], width="stretch", hide_index=True)

elif menu == "Alertas E-mail":
    pagina_alertas_email(df_base, nome_arquivo_exibicao, cliente_nome)

elif menu == "Relatórios":
    st.markdown("<div class='section-title'>Relatório Executivo</div>", unsafe_allow_html=True)
    st.write(f"Total de atividades avaliadas: **{total}**")
    st.write(f"Conclusão geral: **{conclusao_geral:.1f}%**")
    st.write(f"Atividades em atraso: **{qtd_atraso}**")
    st.write(f"Atividades em atenção: **{qtd_atencao}**")
    st.dataframe(resumo, width="stretch", hide_index=True)

elif menu == "Exportar":
    st.markdown("<div class='section-title'>Exportação</div>", unsafe_allow_html=True)
    excel_bytes = gerar_excel(df, resumo)
    st.download_button(
        "⬇️ Baixar relatório Excel",
        data=excel_bytes,
        file_name=f"Relatorio_PMO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    csv = df.to_csv(index=False, sep=";").encode("utf-8-sig")
    st.download_button("⬇️ Baixar atividades CSV", data=csv, file_name="atividades_pmo.csv", mime="text/csv")

elif menu == "Configurações":
    st.markdown("<div class='section-title'>Configurações do Sistema</div>", unsafe_allow_html=True)
    st.info("Para mudar colunas, linhas ou aba padrão, ajuste as constantes no início do arquivo Python.")
    st.code(
        f"ABA_PADRAO = '{ABA_PADRAO}'\nLINHA_INICIAL = {LINHA_INICIAL}\nLINHA_FINAL = {LINHA_FINAL}\nCOL_RESPONSAVEL = '{COL_RESPONSAVEL}'\nCOL_DATA_ALVO = '{COL_DATA_ALVO}'\nCOL_DATA_CONCLUSAO = '{COL_DATA_CONCLUSAO}'"
    )

elif menu == "Administração de Usuários":
    pagina_administrar_usuarios()

# =========================================================
# TABELA FINAL COM CORES FIXAS
# =========================================================
if menu != "Administração de Usuários":
    st.divider()
    st.markdown("<div class='section-title'>Atividades</div>", unsafe_allow_html=True)

    colunas_exibir = ["Item", "Fase", "Processo", "Atividade", "Responsável", "E-mail Responsável", "Data Alvo", "Data Entrega", COL_DEPENDENCIA, "Dias para Vencer", "Status"]

    for col in ["Data Alvo", "Data Entrega"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    def colorir_status(row):
        cor = CORES_STATUS.get(row["Status"], "#FFFFFF")
        texto = "#FFFFFF" if row["Status"] in ["Em Atraso", "Concluida atrasada", "Sem data alvo"] else "#111827"
        return [f"background-color: {cor}; color: {texto}; font-weight: 600" if c == "Status" else "" for c in row.index]

    st.dataframe(
        df[colunas_exibir].style.apply(colorir_status, axis=1).format({
            "Data Alvo": lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) else "",
            "Data Entrega": lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) else "",
        }),
        width="stretch",
        hide_index=True
    )

    st.caption(f"{VERSAO_DASHBOARD} | Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Cliente: {cliente_nome if cliente_nome else 'Cliente não informado'} | Fonte: {nome_arquivo_exibicao}")
